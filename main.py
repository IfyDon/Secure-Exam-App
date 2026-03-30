"""
Secure Computer-Based Exam Application
Requirements:
- Fullscreen enforcement
- Tab switch / focus loss detection → exam ends
- Fullscreen exit detection → exam ends
- Right-click disabled
- Copy disabled
- Ctrl shortcuts disabled
- Camera monitoring (small overlay)
- 2-minute timer
- 5 MCQ questions
- Results saved to Excel spreadsheet
"""

import tkinter as tk
from tkinter import messagebox, ttk
import threading
import time
import datetime
import os
import sys

try:
    import cv2
    CAMERA_AVAILABLE = True
except ImportError:
    CAMERA_AVAILABLE = False

try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ─────────────────────────── PYTHON QUESTIONS ───────────────────────────
QUESTIONS = [
    {
        "question": "1. What is the correct way to create a list in Python?",
        "options": ["A. list = (1, 2, 3)", "B. list = [1, 2, 3]", "C. list = {1, 2, 3}", "D. list = <1, 2, 3>"],
        "answer": "B"
    },
    {
        "question": "2. Which of the following is used to define a function in Python?",
        "options": ["A. define", "B. function", "C. def", "D. func"],
        "answer": "C"
    },
    {
        "question": "3. What is the output of print(2 ** 3)?",
        "options": ["A. 6", "B. 8", "C. 9", "D. 5"],
        "answer": "B"
    },
    {
        "question": "4. Which keyword is used to import a module in Python?",
        "options": ["A. include", "B. using", "C. import", "D. require"],
        "answer": "C"
    },
    {
        "question": "5. What does the 'len()' function do?",
        "options": ["A. Returns the length of an object", "B. Converts to lowercase", "C. Rounds a number", "D. Finds the maximum value"],
        "answer": "A"
    },
]

EXAM_DURATION = 120  # seconds
EXCEL_FILE = "exam_results.xlsx"


# ─────────────────────────── EXCEL HELPER ───────────────────────────
def init_excel():
    if not EXCEL_AVAILABLE:
        return
    if os.path.exists(EXCEL_FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Results"

    headers = ["#", "Student Name", "Class", "Registration No.",
               "Score", "Total", "Percentage (%)", "Status", "Date & Time", "Remarks"]
    col_widths = [5, 22, 15, 20, 8, 8, 16, 12, 22, 20]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"
    wb.save(EXCEL_FILE)


def save_result(name, cls, reg, score, total, remarks=""):
    if not EXCEL_AVAILABLE:
        print(f"Result: {name} | {cls} | {reg} | {score}/{total}")
        return
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    row = ws.max_row + 1
    pct = round((score / total) * 100, 1) if total else 0
    status = "PASS" if pct >= 50 else "FAIL"
    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    data = [row - 1, name, cls, reg, score, total, f"{pct}%", status, dt, remarks]
    aligns = [center, left, center, center, center, center, center, center, center, left]

    status_fill = PatternFill("solid", fgColor="C6EFCE" if status == "PASS" else "FFC7CE")
    status_font_color = "276221" if status == "PASS" else "9C0006"

    for col, (val, aln) in enumerate(zip(data, aligns), 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = aln
        cell.border = border
        cell.font = Font(name="Arial", size=10)
        if col == 8:  # Status column
            cell.fill = status_fill
            cell.font = Font(name="Arial", size=10, bold=True, color=status_font_color)

    # Alternate row shading
    if row % 2 == 0:
        row_fill = PatternFill("solid", fgColor="EBF3FB")
        for col in range(1, len(data) + 1):
            c = ws.cell(row=row, column=col)
            if col != 8:
                c.fill = row_fill

    wb.save(EXCEL_FILE)


# ─────────────────────────── REGISTRATION SCREEN (NO PACKAGES PANEL) ───────────────────────────
class RegistrationScreen:
    def __init__(self, root, on_start):
        self.root = root
        self.on_start = on_start
        self.root.title("Secure Exam Portal — Registration")
        self.root.configure(bg="#0D1B2A")

        # Make registration screen fullscreen
        self._make_fullscreen()
        self._build_ui()

    def _make_fullscreen(self):
        """Make the registration window fullscreen"""
        self.root.attributes("-fullscreen", True)
        self.root.attributes("-topmost", True)
        self.root.focus_force()

        # Bind escape key to exit fullscreen (with confirmation)
        self.root.bind("<Escape>", self._exit_fullscreen)

    def _exit_fullscreen(self, event=None):
        """Exit fullscreen with confirmation"""
        if messagebox.askyesno("Exit Fullscreen",
                                "Do you want to exit fullscreen mode?\n\n"
                                "Note: This will close the application.",
                                parent=self.root):
            self.root.destroy()

    def _build_ui(self):
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Main container (fills full screen)
        main_container = tk.Frame(self.root, bg="#0D1B2A")
        main_container.pack(fill="both", expand=True)

        # Canvas and scrollbar
        canvas = tk.Canvas(main_container, bg="#0D1B2A", highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#0D1B2A")

        # Configure scrolling
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        # Create window inside canvas
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Pack scrollbar and canvas
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Update canvas window width when canvas resizes
        def _configure_canvas(event):
            canvas.itemconfig(canvas_window, width=event.width)
        canvas.bind("<Configure>", _configure_canvas)

        # Mouse wheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # --- Build content inside scrollable_frame ---
        # Header
        header_height = int(screen_height * 0.12)
        hdr = tk.Frame(scrollable_frame, bg="#1B3A5C", height=header_height)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)

        header_font_size = int(min(screen_width, screen_height) * 0.035)
        subheader_font_size = int(header_font_size * 0.6)

        tk.Label(hdr, text="🎓 SECURE EXAM PORTAL",
                 font=("Arial", header_font_size, "bold"),
                 bg="#1B3A5C", fg="#F0C040").pack(expand=True)
        tk.Label(hdr, text="Complete your registration to begin",
                 font=("Arial", subheader_font_size),
                 bg="#1B3A5C", fg="#A0B8D0").pack()

        # Form
        form = tk.Frame(scrollable_frame, bg="#0D1B2A", padx=int(screen_width * 0.1), pady=30)
        form.pack(fill="x")

        self.entries = {}
        fields = [
            ("👤  Student Full Name", "name", "e.g. John Adebayo"),
            ("🎓  Class / Level",     "cls",  "e.g. SS3 / 300L"),
            ("📋  Registration No.",  "reg",  "e.g. REG/2024/001"),
        ]

        label_font_size = int(header_font_size * 0.55)
        entry_font_size = int(header_font_size * 0.5)

        for label, key, placeholder in fields:
            tk.Label(form, text=label, font=("Arial", label_font_size, "bold"),
                     bg="#0D1B2A", fg="#C5D8E8").pack(anchor="w", pady=(15, 8))
            entry = tk.Entry(form, font=("Arial", entry_font_size), bg="#162C42", fg="white",
                             insertbackground="white", relief="flat",
                             highlightthickness=2, highlightcolor="#4A90D9",
                             highlightbackground="#2A4A6A")
            entry.pack(fill="x", ipady=15)
            entry.insert(0, placeholder)
            entry.config(fg="#506070")
            entry.bind("<FocusIn>",  lambda e, en=entry, ph=placeholder: self._clear_placeholder(en, ph))
            entry.bind("<FocusOut>", lambda e, en=entry, ph=placeholder: self._restore_placeholder(en, ph))
            self.entries[key] = entry

        # Rules warning
        warn = tk.Frame(form, bg="#251800", relief="flat", pady=15, padx=20)
        warn.pack(fill="x", pady=(25, 0))

        rules_text = ("⚠  EXAM RULES  ⚠\n\n"
                      "•  No tab switching or window changes\n"
                      "•  No right-click functionality\n"
                      "•  No copy/paste or keyboard shortcuts\n"
                      "•  Camera monitoring active throughout\n"
                      "•  2-minute timer for all questions\n"
                      "•  Any violation will end the exam immediately")

        tk.Label(warn, text=rules_text,
                 font=("Arial", label_font_size - 2),
                 bg="#251800", fg="#FFA040", justify="left").pack(anchor="w")

        # START BUTTON
        btn_frame = tk.Frame(scrollable_frame, bg="#0D1B2A", pady=40, padx=int(screen_width * 0.1))
        btn_frame.pack(fill="x")

        btn = tk.Button(
            btn_frame,
            text="▶   START EXAMINATION   ▶",
            font=("Arial", int(header_font_size * 0.7), "bold"),
            bg="#0F8A3C",
            fg="#FFFFFF",
            activebackground="#12AA4A",
            activeforeground="#FFFFFF",
            relief="flat",
            cursor="hand2",
            pady=20,
            bd=0,
            command=self._submit
        )
        btn.pack(fill="x")

        # Hover effect
        btn.bind("<Enter>", lambda e: btn.config(bg="#12AA4A"))
        btn.bind("<Leave>", lambda e: btn.config(bg="#0F8A3C"))

        # Force update scroll region after all widgets are added
        scrollable_frame.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))

    def _clear_placeholder(self, entry, placeholder):
        if entry.get() == placeholder:
            entry.delete(0, "end")
            entry.config(fg="white")

    def _restore_placeholder(self, entry, placeholder):
        if not entry.get():
            entry.insert(0, placeholder)
            entry.config(fg="#607080")

    def _get_val(self, key, placeholder):
        val = self.entries[key].get().strip()
        return "" if val == placeholder else val

    def _submit(self):
        placeholders = {
            "name": "e.g. John Adebayo",
            "cls": "e.g. SS3 / 300L",
            "reg": "e.g. REG/2024/001"
        }
        name = self._get_val("name", placeholders["name"])
        cls = self._get_val("cls", placeholders["cls"])
        reg = self._get_val("reg", placeholders["reg"])

        if not all([name, cls, reg]):
            messagebox.showerror("Incomplete", "Please fill in all fields before starting.", parent=self.root)
            return

        for widget in self.root.winfo_children():
            widget.destroy()

        self.on_start(name, cls, reg)


# ─────────────────────────── EXAM SCREEN ───────────────────────────
class ExamScreen:
    def __init__(self, root, name, cls, reg):
        self.root = root
        self.student_name = name
        self.student_cls = cls
        self.student_reg = reg

        self.answers = {}
        self.current_q = 0
        self.time_left = EXAM_DURATION
        self.exam_ended = False
        self.timer_running = False
        self.focus_guard_active = False

        self.cap = None
        self.cam_label = None

        self._build_fullscreen()
        self._build_ui()
        self._bind_security()
        self._start_timer()
        self._start_camera()

        # Activate focus guard after short delay (allow window to settle)
        self.root.after(1500, self._activate_focus_guard)

    # ── FULLSCREEN ──────────────────────────────────────────────
    def _build_fullscreen(self):
        self.root.title("SECURE EXAM IN PROGRESS — DO NOT SWITCH TABS")
        self.root.configure(bg="#0A0F1E")
        self.root.attributes("-fullscreen", True)
        self.root.attributes("-topmost", True)
        self.root.focus_force()

    # ── SECURITY BINDINGS ────────────────────────────────────────
    def _bind_security(self):
        # Disable right-click
        self.root.bind_all("<Button-3>", lambda e: "break")

        # Disable common Ctrl shortcuts
        blocked_ctrl = ["c", "v", "x", "a", "z", "p", "s", "f", "w", "t", "n", "r", "u"]
        for key in blocked_ctrl:
            self.root.bind_all(f"<Control-{key}>", lambda e: "break")
            self.root.bind_all(f"<Control-{key.upper()}>", lambda e: "break")

        # Disable Alt+F4, Alt+Tab area
        self.root.bind_all("<Alt-F4>", lambda e: "break")
        self.root.bind_all("<Alt-Tab>", lambda e: "break")

        # Disable PrintScreen
        self.root.bind_all("<Print>", lambda e: "break")

        # Detect fullscreen exit
        self.root.bind("<Configure>", self._on_configure)

        # Focus events
        self.root.bind("<FocusOut>", self._on_focus_out)
        self.root.bind("<FocusIn>", self._on_focus_in)

        # Disable Escape key during exam
        self.root.bind("<Escape>", lambda e: "break")

    def _activate_focus_guard(self):
        self.focus_guard_active = True

    def _on_configure(self, event):
        if self.exam_ended:
            return
        # If window is no longer fullscreen-sized
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        if event.widget == self.root:
            if event.width < sw - 50 or event.height < sh - 50:
                self._terminate_exam("⛔ Fullscreen exited! Exam terminated.")

    def _on_focus_out(self, event):
        if not self.focus_guard_active or self.exam_ended:
            return
        if event.widget == self.root:
            self._terminate_exam("⛔ Tab switch / Window change detected! Exam terminated.")

    def _on_focus_in(self, event):
        pass  # Re-focus is OK

    # ── UI BUILD ─────────────────────────────────────────────────
    def _build_ui(self):
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Responsive font sizes
        title_font_size = int(min(screen_width, screen_height) * 0.025)
        info_font_size = int(title_font_size * 0.7)
        timer_font_size = int(title_font_size * 1.1)
        question_font_size = int(title_font_size * 0.9)
        option_font_size = int(title_font_size * 0.8)

        # ── Top bar ──
        topbar_height = int(screen_height * 0.08)
        topbar = tk.Frame(self.root, bg="#0D2137", height=topbar_height)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)

        tk.Label(topbar, text="🔒 SECURE EXAMINATION",
                 font=("Arial", title_font_size, "bold"),
                 bg="#0D2137", fg="#F0C040").pack(side="left", padx=25)

        # Student info
        info_txt = f"👤 {self.student_name}   |   🎓 {self.student_cls}   |   📋 {self.student_reg}"
        tk.Label(topbar, text=info_txt, font=("Arial", info_font_size),
                 bg="#0D2137", fg="#90B8D8").pack(side="left", padx=25)

        # Timer
        self.timer_label = tk.Label(topbar, text="⏱ 02:00",
                                    font=("Arial", timer_font_size, "bold"),
                                    bg="#0D2137", fg="#00FF88")
        self.timer_label.pack(side="right", padx=30)

        # Progress
        prog_frame = tk.Frame(topbar, bg="#0D2137")
        prog_frame.pack(side="right", padx=20)
        self.progress_label = tk.Label(prog_frame, text="Q 1 / 5",
                                       font=("Arial", info_font_size),
                                       bg="#0D2137", fg="#A0C4E0")
        self.progress_label.pack()

        # ── Main content with scrollbar for questions ──
        main_container = tk.Frame(self.root, bg="#0A0F1E")
        main_container.pack(fill="both", expand=True, padx=int(screen_width * 0.08),
                           pady=int(screen_height * 0.05))

        # Create canvas and scrollbar for main content
        canvas = tk.Canvas(main_container, bg="#0A0F1E", highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#0A0F1E")

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Pack scrollbar and canvas
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Mouse wheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Question card
        card = tk.Frame(scrollable_frame, bg="#0E2236", relief="flat",
                        highlightthickness=2, highlightbackground="#1A5080")
        card.pack(fill="both", expand=True, pady=15)

        inner = tk.Frame(card, bg="#0E2236", padx=int(screen_width * 0.05),
                        pady=int(screen_height * 0.04))
        inner.pack(fill="both", expand=True)

        # Question number badge
        q_badge = tk.Frame(inner, bg="#1A5080", padx=20, pady=8)
        q_badge.pack(anchor="w", pady=(0, 20))
        self.q_num_label = tk.Label(q_badge, text="QUESTION 1 OF 5",
                                    font=("Arial", info_font_size, "bold"),
                                    bg="#1A5080", fg="#A8D4F0")
        self.q_num_label.pack()

        # Question text
        self.q_label = tk.Label(inner, text="", font=("Arial", question_font_size, "bold"),
                                bg="#0E2236", fg="white", wraplength=screen_width - 300,
                                justify="left", anchor="w")
        self.q_label.pack(anchor="w", pady=(0, 30))

        # Separator
        tk.Frame(inner, bg="#1A4060", height=2).pack(fill="x", pady=(0, 30))

        # Options
        self.option_var = tk.StringVar(value="")
        self.option_buttons = []

        options_frame = tk.Frame(inner, bg="#0E2236")
        options_frame.pack(fill="x")

        for i in range(4):
            btn_frame = tk.Frame(options_frame, bg="#0E2236", pady=8)
            btn_frame.pack(fill="x")
            rb = tk.Radiobutton(
                btn_frame,
                text="",
                variable=self.option_var,
                value=str(i),
                font=("Arial", option_font_size),
                bg="#122A42",
                fg="white",
                selectcolor="#1A5C8A",
                activebackground="#1A3A55",
                activeforeground="white",
                indicatoron=False,
                relief="flat",
                anchor="w",
                padx=30,
                pady=18,
                cursor="hand2",
                highlightthickness=0,
                bd=0
            )
            rb.pack(fill="x")
            self.option_buttons.append(rb)

        # ── Navigation ──
        nav_height = int(screen_height * 0.08)
        nav = tk.Frame(self.root, bg="#0D2137", height=nav_height)
        nav.pack(fill="x", side="bottom")
        nav.pack_propagate(False)

        button_font_size = int(info_font_size * 1.1)

        self.prev_btn = tk.Button(nav, text="◀  Previous",
                                  font=("Arial", button_font_size, "bold"),
                                  bg="#1A3A5C", fg="white", activebackground="#2A5080",
                                  relief="flat", padx=30, pady=12, cursor="hand2",
                                  command=self._prev_question)
        self.prev_btn.pack(side="left", padx=30)

        self.next_btn = tk.Button(nav, text="Next  ▶",
                                  font=("Arial", button_font_size, "bold"),
                                  bg="#1A5C3A", fg="white", activebackground="#2A8050",
                                  relief="flat", padx=30, pady=12, cursor="hand2",
                                  command=self._next_question)
        self.next_btn.pack(side="right", padx=30)

        self.submit_btn = tk.Button(nav, text="✅  SUBMIT EXAM",
                                    font=("Arial", button_font_size, "bold"),
                                    bg="#8B0000", fg="white", activebackground="#AA0000",
                                    relief="flat", padx=40, pady=12, cursor="hand2",
                                    command=self._submit_exam)
        self.submit_btn.pack(side="right", padx=20)

        # Camera overlay placeholder
        cam_width = int(screen_width * 0.12)
        cam_height = int(screen_height * 0.12)
        self.cam_frame = tk.Frame(self.root, bg="#000000", width=cam_width, height=cam_height,
                                  highlightthickness=2, highlightbackground="#F0C040")
        self.cam_frame.place(relx=1.0, rely=1.0, anchor="se", x=-15, y=-int(screen_height * 0.05))
        self.cam_frame.pack_propagate(False)
        self.cam_label = tk.Label(self.cam_frame, bg="#000000",
                                  text="📷 No Camera", fg="#607080",
                                  font=("Arial", int(info_font_size * 0.8)))
        self.cam_label.pack(expand=True)

        monitor_font = int(info_font_size * 0.6)
        tk.Label(self.root, text="📷 MONITORING", font=("Arial", monitor_font, "bold"),
                 bg="#F0C040", fg="#000000").place(relx=1.0, rely=1.0, anchor="se",
                                                   x=-15, y=-int(screen_height * 0.05) - monitor_font - 5)

        self._show_question(0)

    # ── QUESTION NAVIGATION ──────────────────────────────────────
    def _show_question(self, idx):
        q = QUESTIONS[idx]
        self.q_num_label.config(text=f"QUESTION {idx+1} OF {len(QUESTIONS)}")
        self.progress_label.config(text=f"Q {idx+1} / {len(QUESTIONS)}")
        self.q_label.config(text=q["question"])
        self.option_var.set(self.answers.get(idx, ""))

        for i, (btn, opt) in enumerate(zip(self.option_buttons, q["options"])):
            btn.config(text=opt, value=str(i))

        self.prev_btn.config(state="normal" if idx > 0 else "disabled")

        if idx == len(QUESTIONS) - 1:
            self.next_btn.config(state="disabled")
        else:
            self.next_btn.config(state="normal")

    def _save_current_answer(self):
        val = self.option_var.get()
        if val != "":
            self.answers[self.current_q] = val

    def _next_question(self):
        self._save_current_answer()
        if self.current_q < len(QUESTIONS) - 1:
            self.current_q += 1
            self._show_question(self.current_q)

    def _prev_question(self):
        self._save_current_answer()
        if self.current_q > 0:
            self.current_q -= 1
            self._show_question(self.current_q)

    # ── TIMER ────────────────────────────────────────────────────
    def _start_timer(self):
        self.timer_running = True
        self._tick()

    def _tick(self):
        if self.exam_ended:
            return
        mins = self.time_left // 60
        secs = self.time_left % 60
        color = "#FF4444" if self.time_left <= 30 else "#00FF88"
        self.timer_label.config(text=f"⏱ {mins:02d}:{secs:02d}", fg=color)

        if self.time_left <= 0:
            self._submit_exam(timeout=True)
            return

        self.time_left -= 1
        self.root.after(1000, self._tick)

    # ── CAMERA ───────────────────────────────────────────────────
    def _start_camera(self):
        if not CAMERA_AVAILABLE or not PIL_AVAILABLE:
            return
        self.cap = cv2.VideoCapture(0)
        if self.cap.isOpened():
            self._update_camera()

    def _update_camera(self):
        if self.exam_ended or not self.cap:
            return
        ret, frame = self.cap.read()
        if ret:
            # Get current camera frame dimensions
            cam_width = self.cam_frame.winfo_width()
            cam_height = self.cam_frame.winfo_height()
            if cam_width > 10 and cam_height > 10:
                frame = cv2.flip(frame, 1)
                frame = cv2.resize(frame, (cam_width, cam_height))
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                img = Image.fromarray(frame_rgb)
                imgtk = ImageTk.PhotoImage(image=img)
                self.cam_label.config(image=imgtk, text="")
                self.cam_label.image = imgtk
        self.root.after(50, self._update_camera)

    def _stop_camera(self):
        if self.cap:
            self.cap.release()
            self.cap = None

    # ── EXAM END ─────────────────────────────────────────────────
    def _terminate_exam(self, reason):
        if self.exam_ended:
            return
        self.exam_ended = True
        self.timer_running = False
        self.focus_guard_active = False
        self._stop_camera()

        score = self._calculate_score()
        save_result(self.student_name, self.student_cls, self.student_reg,
                    score, len(QUESTIONS), remarks=reason)
        self._show_terminated_screen(reason, score)

    def _submit_exam(self, timeout=False):
        if self.exam_ended:
            return
        self._save_current_answer()

        if not timeout:
            answered = len(self.answers)
            if answered < len(QUESTIONS):
                remaining = len(QUESTIONS) - answered
                if not messagebox.askyesno(
                    "Unanswered Questions",
                    f"You have {remaining} unanswered question(s).\nSubmit anyway?",
                    parent=self.root
                ):
                    return

        self.exam_ended = True
        self.timer_running = False
        self.focus_guard_active = False
        self._stop_camera()

        score = self._calculate_score()
        reason = "Time Expired" if timeout else "Student Submitted"
        save_result(self.student_name, self.student_cls, self.student_reg,
                    score, len(QUESTIONS), remarks=reason)
        self._show_result_screen(score, timeout)

    def _calculate_score(self):
        score = 0
        for idx, q in enumerate(QUESTIONS):
            if idx in self.answers:
                selected_idx = int(self.answers[idx])
                selected_letter = ["A", "B", "C", "D"][selected_idx]
                if selected_letter == q["answer"]:
                    score += 1
        return score

    # ── RESULT SCREENS ───────────────────────────────────────────
    def _clear_screen(self):
        self.focus_guard_active = False
        for w in self.root.winfo_children():
            w.destroy()

    def _show_terminated_screen(self, reason, score):
        self._clear_screen()
        self.root.configure(bg="#1A0000")

        # Keep fullscreen
        self.root.attributes("-fullscreen", True)

        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        frame = tk.Frame(self.root, bg="#1A0000")
        frame.place(relx=0.5, rely=0.5, anchor="center")

        title_font = int(min(screen_width, screen_height) * 0.06)
        text_font = int(title_font * 0.4)

        tk.Label(frame, text="🚫", font=("Arial", int(title_font * 1.5)),
                 bg="#1A0000", fg="#FF2222").pack()
        tk.Label(frame, text="EXAMINATION TERMINATED",
                 font=("Arial", title_font, "bold"),
                 bg="#1A0000", fg="#FF4444").pack(pady=(15, 8))
        tk.Label(frame, text=reason, font=("Arial", text_font + 4),
                 bg="#1A0000", fg="#FF8888").pack(pady=8)
        tk.Label(frame, text=f"Score recorded: {score} / {len(QUESTIONS)}",
                 font=("Arial", text_font), bg="#1A0000", fg="#AAAAAA").pack(pady=12)

        if EXCEL_AVAILABLE:
            tk.Label(frame, text=f"✅ Result saved to: {EXCEL_FILE}",
                     font=("Arial", text_font - 2), bg="#1A0000", fg="#888888").pack(pady=5)

        tk.Button(frame, text="Close", font=("Arial", text_font + 2, "bold"),
                  bg="#440000", fg="white", relief="flat", padx=40, pady=12,
                  command=self.root.destroy, cursor="hand2").pack(pady=25)

    def _show_result_screen(self, score, timeout):
        self._clear_screen()
        total = len(QUESTIONS)
        pct = round((score / total) * 100)
        passed = pct >= 50

        bg = "#001A00" if passed else "#1A0A00"
        self.root.configure(bg=bg)

        # Keep fullscreen
        self.root.attributes("-fullscreen", True)

        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        frame = tk.Frame(self.root, bg=bg)
        frame.place(relx=0.5, rely=0.5, anchor="center")

        title_font = int(min(screen_width, screen_height) * 0.06)
        text_font = int(title_font * 0.4)

        tk.Label(frame, text="✅" if passed else "❌",
                 font=("Arial", int(title_font * 1.5)),
                 bg=bg, fg="#00FF88" if passed else "#FF8844").pack()
        tk.Label(frame, text="EXAM COMPLETE", font=("Arial", title_font, "bold"),
                 bg=bg, fg="white").pack(pady=(15, 8))

        # Score circle
        score_frame = tk.Frame(frame, bg="#0A2A0A" if passed else "#2A1A00",
                               relief="flat", padx=int(screen_width * 0.05),
                               pady=int(screen_height * 0.03),
                               highlightthickness=3,
                               highlightbackground="#00AA44" if passed else "#AA6600")
        score_frame.pack(pady=20)

        tk.Label(score_frame, text=f"{score} / {total}",
                 font=("Arial", int(title_font * 0.8), "bold"),
                 bg=score_frame["bg"], fg="#00FF88" if passed else "#FFAA44").pack()
        tk.Label(score_frame, text=f"{pct}%  —  {'PASS ✓' if passed else 'FAIL ✗'}",
                 font=("Arial", text_font + 4, "bold"),
                 bg=score_frame["bg"], fg="#80FF80" if passed else "#FFBB66").pack()

        # Student info
        info = f"👤 {self.student_name}   |   🎓 {self.student_cls}   |   📋 {self.student_reg}"
        tk.Label(frame, text=info, font=("Arial", text_font),
                 bg=bg, fg="#A0B8C8").pack(pady=10)

        if timeout:
            tk.Label(frame, text="⏱ Exam ended: Time expired",
                     font=("Arial", text_font - 1), bg=bg, fg="#FFAA44").pack()

        if EXCEL_AVAILABLE:
            tk.Label(frame, text=f"✅ Result saved to: {EXCEL_FILE}",
                     font=("Arial", text_font - 2), bg=bg, fg="#607080").pack(pady=6)

        tk.Button(frame, text="Close", font=("Arial", text_font + 2, "bold"),
                  bg="#0A3A1A" if passed else "#3A1A00", fg="white",
                  relief="flat", padx=40, pady=12,
                  command=self.root.destroy, cursor="hand2").pack(pady=25)


# ─────────────────────────── MAIN ───────────────────────────────
def main():
    # Check dependencies
    missing = []
    if not EXCEL_AVAILABLE:
        missing.append("openpyxl")
    if not CAMERA_AVAILABLE:
        missing.append("opencv-python")
    if not PIL_AVAILABLE:
        missing.append("Pillow")

    if missing:
        print(f"[WARN] Missing packages: {', '.join(missing)}")
        print(f"       Install with: pip install {' '.join(missing)}")
        print("       App will run with reduced functionality.\n")

    init_excel()

    root = tk.Tk()
    root.overrideredirect(False)  # Keep window decorations

    def start_exam(name, cls, reg):
        ExamScreen(root, name, cls, reg)

    RegistrationScreen(root, start_exam)
    root.mainloop()


if __name__ == "__main__":
    main()